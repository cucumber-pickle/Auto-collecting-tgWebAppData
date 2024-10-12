import os
import subprocess
import time
from pywinauto import Application
import openpyxl
import pyperclip
import pyautogui
from src.cucumber import log, mrh, pth, hju, kng, bru, htm
from global_config import base_dir, file_path, console_image_paths, seach_image_paths, close_image_paths


def load_accounts_from_file(file_path):
    """Loads accounts from a text file and returns them as a list."""
    with open(file_path, 'r') as file:
        # Read lines from the file and convert them to integers
        accounts = [int(line.strip()) for line in file if line.strip().isdigit()]
    return accounts

def load_bots_from_file(file_path):
    """Loads bots from a text file and returns them as a list."""
    with open(file_path, 'r') as file:
        # Read lines from the file and remove extra spaces
        bots = [line.strip() for line in file if line.strip()]
    return bots

def launch_telegram(account_num):
    account_folder = os.path.join(base_dir, str(account_num))
    try:
        exe_file = os.path.join(account_folder, f"{account_num}.exe")
        subprocess.Popen(exe_file)
        log(f"Launching Telegram for account {account_num} from {kng}{exe_file}")
    except:
        exe_file = os.path.join(account_folder, f"Telegram.exe")
        subprocess.Popen(exe_file)
        log(f"Launching Telegram for account {account_num} from {exe_file}...")

    time.sleep(10)  # Delay to wait for Telegram to launch
    log(hju + "Telegram launched.")

    return Application().connect(path=exe_file), exe_file

def close_telegram(account_num, exe_file):
    log(f"Closing Telegram for account {account_num}...")
    subprocess.call(f'taskkill /F /IM {os.path.basename(exe_file)}', shell=True)
    time.sleep(4)  # Delay to close

def write_to_excel(account_num, bot_name, row):
    # Open the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get all values from the first row
    first_row_values = [cell.value for cell in sheet[1]]

    # Check if bot_name exists in the first row
    if bot_name in first_row_values:
        # Find the column index for bot_name
        col_index = first_row_values.index(bot_name) + 1  # +1 for correct indexing in openpyxl
    else:
        # Add the bot name to the first row and get the index of the new column
        col_index = len(first_row_values) + 1  # New index for the next column
        sheet.cell(row=1, column=col_index, value=bot_name)

    # Copy data from clipboard
    time.sleep(1)  # Small delay before pasting
    tg_webapp_data = pyperclip.paste()

    # Add data to the table
    sheet["A1"] = "account_num"
    sheet[f'A{row}'] = account_num  # Account number
    sheet.cell(row=row, column=col_index, value=tg_webapp_data)  # Data from clipboard

    # Save the file
    workbook.save(file_path)
    log(hju + f"Data for account {account_num} successfully written to {file_path}")

def click_console(bot_name):
    # Search and click on the "Console" tab
    log("Searching for 'Console' tab on the screen...")

    if bot_name == 'agent301bot':
        time.sleep(5)
    else:
        time.sleep(4)

    for image_path in console_image_paths:
        try:
            console_location = pyautogui.locateOnScreen(image_path, confidence=0.9)
            if console_location:  # If the image is found, exit the loop
                log(hju + f"Console found at coordinates: {console_location}")
                break
        except Exception as e:
            log(f'Error while searching for image {image_path}: {e}')

    time.sleep(1)

    if console_location:
        log("Console tab found, clicking on it...")
        pyautogui.click(pyautogui.center(console_location))
        time.sleep(1)
    else:
        log("Failed to find Console tab.")
        return

def kucoin():
    kucoin = r"pic\support\kucoin_button.png"
    time.sleep(9)

    try:
        for _ in range(4):
            kucoin_click = pyautogui.locateOnScreen(kucoin, confidence=0.9)
            time.sleep(1)
            if kucoin_click:
                log("Searching for KuCoin button")
                pyautogui.click(pyautogui.center(kucoin_click), duration=0.5)
                time.sleep(1)
            else:
                log("Failed to find KuCoin button")

        time.sleep(3)

    except Exception as e:
        log(f'Error during KuCoin execution {e}')
        return

def duck_chain():
    duck = r"pic\support\duck_button.png"
    time.sleep(3)

    try:
        try:
            log("Searching for Duck button")
            duck_click = pyautogui.locateOnScreen(duck, confidence=0.9)
            time.sleep(1)
            if duck_click:
                log("Clicking Duck button")
                pyautogui.click(pyautogui.center(duck_click), duration=0.5)
            else:
                log("Failed to find Duck button")

        except Exception as e:
            log(f'Error while clicking Duck button {e}')
            return

        time.sleep(3)

    except Exception as e:
        log(f'Error during Duck execution {e}')
        return

def match_quest():
    match1 = r"pic\support\match_b1.png"
    match2 = r"pic\support\match_b2.png"

    time.sleep(1)

    try:
        try:
            match_click1 = pyautogui.locateOnScreen(match1, confidence=0.7)
            time.sleep(1)
            if match_click1:
                log("Searching for Match1 button")
                pyautogui.click(pyautogui.center(match_click1))
            else:
                log("Failed to find Match1 button")

        except Exception as e:
            log(f'Error while clicking Match1 button {e}')
            return

        time.sleep(5)

        try:
            match_click2 = pyautogui.locateOnScreen(match2, confidence=0.7)
            time.sleep(1)
            if match_click2:
                log("Searching for Match2 button")
                pyautogui.click(pyautogui.center(match_click2))
            else:
                log("Failed to find Match2 button")

        except Exception as e:
            log(f'Error while clicking Match2 image {e}')
            return

        time.sleep(5)

    except Exception as e:
        log(f'Error during Match execution {e}')
        return

def open_bot(app):
    log("Focusing on the Telegram window...")
    main_window = app.top_window()
    main_window.set_focus()

    time.sleep(1.5)

    log("Opening search...")
    main_window.type_keys('^f')  # Ctrl + F

    time.sleep(2.5)

    log(f"Entering group name: ...")

    main_window.type_keys('tapalki_reff_chat')

    time.sleep(2.5)

    log("Pressing down key to select the bot and Enter...")

    main_window.type_keys('{DOWN}')

    time.sleep(1.5)

    main_window.type_keys('{ENTER}')

    time.sleep(1.5)

def close_app():
    try:
        for image_path in close_image_paths:
            try:
                close_location = pyautogui.locateOnScreen(image_path, confidence=0.7)
                if close_location:
                    time.sleep(1)
                    pyautogui.click(pyautogui.center(close_location))
                    log(f'Clicking application close button')
                    time.sleep(1)
                    break
            except:
                log(f'')

    except:
        log('')

def click_seach():
    time.sleep(1)  # Give time for the link to appear

    for image_path in seach_image_paths:
        try:
            seach_location = pyautogui.locateOnScreen(image_path, confidence=0.8)
            time.sleep(1)
            if seach_location:  # If the image is found, exit the loop
                break
        except Exception as e:
            log(f'Error while searching for image {image_path}: {e}')

    pyautogui.click(pyautogui.center(seach_location))
    time.sleep(1)  # Delay before pressing Enter