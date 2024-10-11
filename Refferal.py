import os
import subprocess
import time
import random
from pywinauto import Application
from pywinauto.keyboard import send_keys
import pyautogui
import openpyxl
import pyperclip
import keyboard  # New import for keyboard interaction
from colorama import Fore, Style
from datetime import datetime
from global_config import close_image_paths, seach_image_paths, base_dir, file_path, console_image_paths, codes

pth = Fore.LIGHTWHITE_EX
reset = Style.RESET_ALL
putih = Fore.LIGHTWHITE_EX
hitam = Fore.LIGHTBLACK_EX


def countdown_timer(seconds):
    while seconds:
        m, s = divmod(seconds, 60)
        h, m = divmod(m, 60)
        h = str(h).zfill(2)
        m = str(m).zfill(2)
        s = str(s).zfill(2)
        print(f"{pth}please wait until {h}:{m}:{s} ", flush=True, end="\r")
        seconds -= 1
        time.sleep(1)
    print(f"{pth}please wait until {h}:{m}:{s} ", flush=True, end="\r")

def log(message):
    now = datetime.now().isoformat(" ").split(".")[0]
    print(f"{hitam}[{now}]{putih} {message}{reset}")


# Функция для запуска Telegram Portable
def launch_telegram(account_num):

    account_folder = os.path.join(base_dir, str(account_num))
    try:
        exe_file = os.path.join(account_folder, f"{account_num}.exe")
        subprocess.Popen(exe_file)
        log(f"Запуск Telegram для аккаунта {account_num} из {exe_file}...")
    except:
        exe_file = os.path.join(account_folder, f"Telegram.exe")
        subprocess.Popen(exe_file)
        log(f"Запуск Telegram для аккаунта {account_num} из {exe_file}...")

    time.sleep(8)  # Задержка для ожидания запуска Telegram
    log("Telegram запущен.")

    return Application().connect(path=exe_file)

# Функция для закрытия Telegram
def close_telegram(account_num):
    log(f"Закрываем Telegram для аккаунта {account_num}...")
    exe_file = os.path.join(base_dir, str(account_num), f"{account_num}.exe")
    subprocess.call(f'taskkill /F /IM {os.path.basename(exe_file)}', shell=True)
    time.sleep(3)  # Задержка для закрытия

# Функция для поиска и нажатия на ссылку
def click_link(bot_name):  # Даем время для отображения ссылки

    link= fr"pic\bots\{bot_name}.png"
    try:
        link_location = pyautogui.locateOnScreen(link, confidence=0.7)
        time.sleep(1)

    except Exception as e:
        log(f'Ошибка при поиске изображения {bot_name}: {e}')

    if link_location:
        log("Ссылка найдена. Нажимаем на неё...")
        pyautogui.click(pyautogui.center(link_location))
        time.sleep(2)  # Задержка перед нажатием Enter
        send_keys('{ENTER}')  # Нажимаем Enter через 3 секунды
    else:
        log("Не удалось найти ссылку.")

def click_seach():
    time.sleep(1)  # Даем время для отображения ссылки

    for image_path in seach_image_paths:
        try:
            seach_location = pyautogui.locateOnScreen(image_path, confidence=0.8)
            time.sleep(1)
            if seach_location:  # Если изображение найдено, выходим из цикла
                break
        except Exception as e:
            log(f'Ошибка при поиске изображения {image_path}: {e}')


    pyautogui.click(pyautogui.center(seach_location))
    time.sleep(1)  # Задержка перед нажатием Enter

def open_bot(app):
    log("Фокусируемся на окне Telegram...")
    main_window = app.top_window()
    main_window.set_focus()
    time.sleep(1.5)

    log("Открываем поиск...")
    main_window.type_keys('^f')  # Ctrl + F
    time.sleep(2.5)

    log(f"Вводим название группы: ...")
    main_window.type_keys('tapalki_reff_chat')
    time.sleep(2.5)

    log("Нажимаем клавишу вниз для выбора бота и Enter...")
    main_window.type_keys('{DOWN}')
    time.sleep(1.5)
    main_window.type_keys('{ENTER}')
    time.sleep(1.5)
# Функция для взаимодействия с ботом

def interact_with_bot(app, bot_name, code, account_num, row):
    log("Фокусируемся на окне Telegram...")
    main_window = app.top_window()
    main_window.set_focus()

    log(f"Нажимаем кнопку поиск")
    click_seach()

    log(f"Вводим название бота: ...")
    main_window.type_keys(bot_name)
    time.sleep(2)


    log("Листаем вниз от 1 до 3 раз")
    num_down_presses = random.randint(1, 3)
    for _ in range(num_down_presses):
        main_window.type_keys('{DOWN}')
        time.sleep(0.5)
    main_window.type_keys('{ENTER}')
    time.sleep(5)

    # Нажимаем на ссылку для запуска мини-приложения
    click_link(bot_name)

    time.sleep(2)

    # Открываем панель разработчика (нажимаем F12)
    log("Нажимаем F12 для открытия панели разработчика...")
    send_keys('{F12}')

    # Поиск и нажатие на вкладку "Консоль"
    log("Ищем вкладку 'Консоль' на экране...")

    if bot_name == 'agent301bot':
        time.sleep(5)
    else:
        time.sleep(4)

    for image_path in console_image_paths:
        try:
            console_location = pyautogui.locateOnScreen(image_path, confidence=0.8)
            time.sleep(1)
            if console_location:  # Если изображение найдено, выходим из цикла
                break
        except Exception as e:
            log(f'Ошибка при поиске изображения {image_path}: {e}')

    time.sleep(1)

    if console_location:
        log("Вкладка 'Консоль' найдена, нажимаем на неё...")
        pyautogui.click(pyautogui.center(console_location))
        time.sleep(1)
    else:
        log("Не удалось найти вкладку 'Консоль'.")
        return

    # Переключение на английскую раскладку
    keyboard.write('`')  # Simulating the ` key for switching console

    # Нажимаем Ctrl+` для активации командной строки
    log("Активируем командную строку с помощью Ctrl+`...")
    keyboard.press_and_release('ctrl+`')
    time.sleep(1)

    # Вводим команду в консоль с задержками для скобок
    # Отправляем строку с кодом
    send_keys(code)
    time.sleep(1)

    # Нажимаем Enter
    send_keys('{ENTER}')
    time.sleep(1)

    # Копируем данные из буфера обмена и записываем в эксель
    write_to_excel(account_num, bot_name, row)

    # Функция для записи данных в эксель
    log("закрываем 2 окна")
    keyboard.press_and_release('alt+F4')
    time.sleep(0.5)

    if bot_name == 'matchquest':
        MatchQuest()

    if bot_name == 'duckchain':
        DuckChain()

    if bot_name == 'kucoin':
        Kucoin()

    keyboard.press_and_release('alt+F4')
    time.sleep(0.5)

    try:
        for image_path in close_image_paths:
            try:
                close_location = pyautogui.locateOnScreen(image_path, confidence=0.7)
                if close_location:
                    time.sleep(1)
                    pyautogui.click(pyautogui.center(close_location))
                    log(f'Нажимаем кнопку закрытия приложения')
                    time.sleep(1)
                    break
            except:
                log(f'')
    except:
        log('')


    keyboard.press_and_release('Esc')
    time.sleep(0.5)


def Kucoin():
    kucoin = r"pic\support\kucoin_button.png"
    time.sleep(9)
    try:
        for _ in range(4):
            kucoin_click = pyautogui.locateOnScreen(kucoin, confidence=0.9)
            time.sleep(1)
            if kucoin_click:
                log("Ищем кнопку Кукоин")
                pyautogui.click(pyautogui.center(kucoin_click), duration=0.5)
                time.sleep(1)
            else:
                log("Не удалось найти кнопку Кукоин")
        time.sleep(3)
    except Exception as e:
        log(f'Ошибка при выполнении Кукоин {e}')
        return

def DuckChain():
    duck = r"pic\support\duck_button.png"
    time.sleep(3)
    try:
        try:
            log("Ищем кнопку Дак")
            duck_click = pyautogui.locateOnScreen(duck, confidence=0.9)
            time.sleep(1)
            if duck_click:
                log("нажимаем кнопку Дак")
                pyautogui.click(pyautogui.center(duck_click), duration=0.5)

            else:
                log("Не удалось найти кнопку Дак")

        except Exception as e:
            log(f'Ошибка при нажатии кнопки Дак {e}')
            return

        time.sleep(3)

    except Exception as e:
        log(f'Ошибка при выполении Дак {e}')
        return

def MatchQuest():
    match1 = r"pic\support\match_b1.png"
    match2 = r"pic\support\match_b2.png"
    time.sleep(1)
    try:
        try:
            match_click1 = pyautogui.locateOnScreen(match1, confidence=0.7)
            time.sleep(1)
            if match_click1 :
                log("Ищем кнопку Матч1")
                pyautogui.click(pyautogui.center(match_click1))
            else:
                log("Не удалось найти кнопку Матч1")

        except Exception as e:
            log(f'Ошибка при нажатии кнопки Матч1 {e}')
            return

        time.sleep(5)

        try:
            match_click2 = pyautogui.locateOnScreen(match2, confidence=0.7)
            time.sleep(1)
            if match_click2 :
                log("Ищем кнопку Матч2")
                pyautogui.click(pyautogui.center(match_click2))
            else:
                log("Не удалось найти кнопку Матч2")

        except Exception as e:
            log(f'Ошибка при нажатии изображения Матч2 {e}')
            return
        time.sleep(5)
    except Exception as e:
        log(f'Ошибка при выполении Матч {e}')
        return


def write_to_excel(account_num, bot_name, row):
    # Открываем эксель файл
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Получаем все значения из первой строки
    first_row_values = [cell.value for cell in sheet[1]]


    # Проверяем, существует ли bot_name в первой строке
    if bot_name in first_row_values:
        # Находим индекс столбца для bot_name
        col_index = first_row_values.index(bot_name) + 1  # +1 для корректного индексации в openpyxl
    else:
        # Добавляем название бота в первую строку и получаем индекс нового столбца
        col_index = len(first_row_values) + 1  # Новый индекс для следующего столбца
        sheet.cell(row=1, column=col_index, value=bot_name)

    # Копируем данные из буфера обмена
    time.sleep(1)  # Небольшая задержка перед вставкой
    tg_webapp_data = pyperclip.paste()

    # Добавляем данные в таблицу
    sheet["A1"] = "account_num"
    sheet[f'A{row}'] = account_num  # Номер аккаунта
    sheet.cell(row=row, column=col_index, value=tg_webapp_data)  # Данные из буфера обмена

    # Сохраняем файл
    workbook.save(file_path)
    log(f"Данные для аккаунта {account_num} успешно записаны в {file_path}")

# Основная функция для работы с несколькими аккаунтами
def process_accounts(account, bots_list, row):
    try:
        app = launch_telegram(account)
        open_bot(app)
        for bot in bots_list:
            # Get the specific code for the bot, or use the default if not found
            code = codes.get(bot, "copy+9Telegram.WebApp.initData+0")
            interact_with_bot(app, bot, code, account, row)
    except Exception as e:
        log(f"Ошибка при взаимодействии с аккаунтом {account}: {e}")
        open("error.log", "a", encoding="utf-8").write(
            f"Ошибка при взаимодействии с аккаунтом {account}: {e} \n")

    finally:
        close_telegram(account)


def load_accounts_from_file(file_path):
    """Загружает аккаунты из текстового файла и возвращает их в виде списка."""
    with open(file_path, 'r') as file:
        # Читаем строки из файла и преобразуем их в целые числа
        accounts = [int(line.strip()) for line in file if line.strip().isdigit()]
    return accounts

def load_bots_from_file(file_path):
    """Загружает ботов из текстового файла и возвращает их в виде списка."""
    with open(file_path, 'r') as file:
        # Читаем строки из файла и удаляем лишние пробелы
        bots = [line.strip() for line in file if line.strip()]
    return bots

# Загрузка аккаунтов из файла acc.txt
accounts_list = load_accounts_from_file('acc.txt')
bots_list = load_bots_from_file('bots.txt')


log(f'Начинаем сбор квери для ботов: {bots_list}')

row = 1

for account in accounts_list:
    row += 1
    print(row)
    process_accounts(account, bots_list, row)
